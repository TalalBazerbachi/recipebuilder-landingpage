#!/usr/bin/env python3
"""
Unified ingredient scraper for FoodLabelMaker API.

Discovers ingredients via search, then fetches full details (nutrients, allergens)
for each one. Both steps are resumable — safe to Ctrl+C anytime.

Usage:
    python3 scrape_ingredients.py          # Scrape + fetch details + export Excel
    python3 scrape_ingredients.py --excel  # Just export saved data to Excel

Files:
    scrape_ingredients.json  — ingredient list from search API
    scrape_progress.json     — search scraper bookmark (phase/query)
    ingredient_details.json  — full details per ingredient (nutrients etc.)
    ingredients.xlsx         — final Excel output
"""
import json
import time
import string
import sys
import os
import signal
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(line_buffering=True)

BASE = os.path.dirname(os.path.abspath(__file__))
PROGRESS_FILE = os.path.join(BASE, "scrape_progress.json")
INGREDIENTS_FILE = os.path.join(BASE, "scrape_ingredients.json")
DETAILS_FILE = os.path.join(BASE, "ingredient_details.json")
EXCEL_FILE = os.path.join(BASE, "ingredients.xlsx")

AUTH_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJpYXQiOjE3NzIxOTMxNDIsImp0aSI6InRva2VuIiwiaXNzIjoiRm9vZExhYmVsTWFrZXIiLCJuYmYiOjE3NzIxOTMxNDIsImV4cCI6MTc3Mjc5Nzk0MiwiZGF0YSI6eyJ1c2VySWQiOjEzODYyNSwiaXNNYXN0ZXJVc2VyIjp0cnVlLCJyb2xlTmFtZSI6InVzZXIiLCJyb2xlSWQiOjEsImFjY291bnRJZCI6MTM4NzA1LCJhY2NvdW50VHlwZUlkIjoyOX19.aUI81wJmHHHsRUx3jBnwjaVxIgzBpnztIPcGpMy-2ampfoW8qEOtMqJLIi59F-l1HqphzQTVHi62ocrBm4X3OQ"

SEARCH_URL = "https://api.foodlabelmaker.com/api/ingredients/search"
DETAIL_URL = "https://api.foodlabelmaker.com/api/ingredients"
SEARCH_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "authorization": AUTH_TOKEN,
    "content-type": "application/json",
    "x-locale": "en",
}
DETAIL_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "authorization": AUTH_TOKEN,
    "x-locale": "en",
}
FILTER = 'visibility = "public" AND dataset != "usda_branded_foods" AND dataset != "canadian_foods"'

# ---------------------------------------------------------------------------
# Global state
# ---------------------------------------------------------------------------
all_ingredients = {}   # id (int) -> search result dict
details = {}           # str(id) -> full detail dict or None
dirty_ingredients = False
dirty_details = False

def save_all():
    global dirty_ingredients, dirty_details
    if dirty_ingredients:
        with open(INGREDIENTS_FILE, "w") as f:
            json.dump(all_ingredients, f, ensure_ascii=False)
        dirty_ingredients = False
    if dirty_details:
        with open(DETAILS_FILE, "w") as f:
            json.dump(details, f, ensure_ascii=False)
        dirty_details = False

def handle_signal(signum, frame):
    print(f"\n\nInterrupted! Saving progress...")
    save_all()
    ing_count = len(all_ingredients)
    det_count = sum(1 for v in details.values() if v is not None)
    print(f"Saved: {ing_count} ingredients listed, {det_count} with full details.")
    print("Run again to resume.")
    sys.exit(0)

signal.signal(signal.SIGINT, handle_signal)
signal.signal(signal.SIGTERM, handle_signal)

# ---------------------------------------------------------------------------
# Search API
# ---------------------------------------------------------------------------
def search(query, retries=3):
    payload = {"query": query, "filter": FILTER}
    for attempt in range(retries):
        try:
            r = requests.post(SEARCH_URL, json=payload, headers=SEARCH_HEADERS, timeout=15)
            if r.status_code == 200:
                data = r.json().get("data", {})
                return data.get("hits", []), data.get("nbHits", 0)
        except:
            pass
        time.sleep(1)
    return [], 0

def save_search_progress(phase, last_query, need_drill_3=None, need_drill_4=None):
    with open(PROGRESS_FILE, "w") as f:
        json.dump({
            "phase": phase,
            "last_query": last_query,
            "total": len(all_ingredients),
            "need_drill_3": need_drill_3 or [],
            "need_drill_4": need_drill_4 or [],
        }, f)

def load_search_progress():
    if os.path.exists(PROGRESS_FILE) and os.path.exists(INGREDIENTS_FILE):
        with open(PROGRESS_FILE) as f:
            progress = json.load(f)
        with open(INGREDIENTS_FILE) as f:
            data = json.load(f)
        for k, v in data.items():
            all_ingredients[int(k)] = v
        return progress
    return None

# ---------------------------------------------------------------------------
# Detail API
# ---------------------------------------------------------------------------
def fetch_detail(ingredient_id, retries=3):
    for attempt in range(retries):
        try:
            r = requests.get(f"{DETAIL_URL}/{ingredient_id}", headers=DETAIL_HEADERS, timeout=15)
            if r.status_code == 200:
                return r.json().get("data")
            elif r.status_code == 404:
                return None
        except:
            pass
        time.sleep(1)
    return None

def load_details():
    if os.path.exists(DETAILS_FILE):
        with open(DETAILS_FILE) as f:
            details.update(json.load(f))

def fetch_details_for_new():
    """Fetch details for any ingredients that don't have details yet."""
    global dirty_details
    missing = [iid for iid in all_ingredients if str(iid) not in details]
    if not missing:
        return
    print(f"\n--- Fetching details for {len(missing)} new ingredients ---")
    for idx, iid in enumerate(missing):
        detail = fetch_detail(iid)
        details[str(iid)] = detail
        dirty_details = True
        if (idx + 1) % 50 == 0:
            save_all()
            fetched = sum(1 for v in details.values() if v is not None)
            print(f"  Details: [{idx+1}/{len(missing)}] — {fetched} total with data — saved")
        time.sleep(0.05)
    save_all()
    fetched = sum(1 for v in details.values() if v is not None)
    print(f"  Details done: {fetched} total with data")

# ---------------------------------------------------------------------------
# Search phases
# ---------------------------------------------------------------------------
def add_hits(hits):
    global dirty_ingredients
    new = 0
    for h in hits:
        if h["id"] not in all_ingredients:
            all_ingredients[h["id"]] = h
            new += 1
            dirty_ingredients = True
    return new

def scrape():
    global dirty_ingredients
    letters = list(string.ascii_lowercase)
    progress = load_search_progress()
    load_details()

    resume_phase = None
    resume_query = None
    need_drill_3 = []
    need_drill_4 = []
    save_counter = 0

    if progress:
        resume_phase = progress["phase"]
        resume_query = progress["last_query"]
        need_drill_3 = progress.get("need_drill_3", [])
        need_drill_4 = progress.get("need_drill_4", [])
        det_count = sum(1 for v in details.values() if v is not None)
        print(f"RESUMING: phase {resume_phase}, query '{resume_query}'")
        print(f"  {len(all_ingredients)} ingredients listed, {det_count} with details")

    if resume_phase == 99:
        print("Search scraping already complete. Checking for missing details...")
        fetch_details_for_new()
        return

    def tick():
        nonlocal save_counter
        save_counter += 1
        if save_counter % 50 == 0:
            save_all()

    # Phase 1: Single letters
    if not resume_phase or resume_phase == 1:
        print("\n=== Phase 1: Single letter queries ===")
        need_drill = []
        skip = resume_phase == 1
        for ch in letters:
            if skip:
                if ch <= (resume_query or ""):
                    hits, _ = search(ch)
                    if len(hits) >= 50:
                        need_drill.append(ch)
                    continue
                skip = False
            hits, nb = search(ch)
            new = add_hits(hits)
            print(f"  '{ch}': {len(hits)} hits, {new} new | Total: {len(all_ingredients)}")
            if len(hits) >= 50:
                need_drill.append(ch)
            save_search_progress(1, ch)
            tick()
            time.sleep(0.1)
        # Fetch details for phase 1 finds
        fetch_details_for_new()
    else:
        need_drill = letters

    # Phase 2: Two-letter combos
    if not resume_phase or resume_phase <= 2:
        if resume_phase and resume_phase > 1:
            need_drill = letters
        print(f"\n=== Phase 2: Two-letter combos ({len(need_drill)} letters) ===")
        skip = resume_phase == 2
        for ch in need_drill:
            for ch2 in letters:
                q = ch + ch2
                if skip:
                    if q <= (resume_query or ""):
                        continue
                    skip = False
                hits, nb = search(q)
                new = add_hits(hits)
                if new > 0 or len(hits) >= 50:
                    print(f"  '{q}': {len(hits)} hits, {new} new | Total: {len(all_ingredients)}")
                if len(hits) >= 50:
                    need_drill_3.append(q)
                save_search_progress(2, q, need_drill_3)
                tick()
                time.sleep(0.08)
        save_all()
        fetch_details_for_new()

    # Phase 3: Three-letter combos
    if not resume_phase or resume_phase <= 3:
        if need_drill_3:
            print(f"\n=== Phase 3: Three-letter combos ({len(need_drill_3)} prefixes) ===")
            skip = resume_phase == 3
            for prefix in need_drill_3:
                for ch3 in letters:
                    q = prefix + ch3
                    if skip:
                        if q <= (resume_query or ""):
                            continue
                        skip = False
                    hits, nb = search(q)
                    new = add_hits(hits)
                    if new > 0 or len(hits) >= 50:
                        print(f"  '{q}': {len(hits)} hits, {new} new | Total: {len(all_ingredients)}")
                    if len(hits) >= 50:
                        need_drill_4.append(q)
                    save_search_progress(3, q, need_drill_3, need_drill_4)
                    tick()
                    time.sleep(0.08)
            save_all()
            fetch_details_for_new()

    # Phase 4: Four-letter combos
    if not resume_phase or resume_phase <= 4:
        if need_drill_4:
            print(f"\n=== Phase 4: Four-letter combos ({len(need_drill_4)} prefixes) ===")
            skip = resume_phase == 4
            for prefix in need_drill_4:
                for ch4 in letters:
                    q = prefix + ch4
                    if skip:
                        if q <= (resume_query or ""):
                            continue
                        skip = False
                    hits, nb = search(q)
                    new = add_hits(hits)
                    if new > 0:
                        print(f"  '{q}': {len(hits)} hits, {new} new | Total: {len(all_ingredients)}")
                    save_search_progress(4, q, need_drill_3, need_drill_4)
                    tick()
                    time.sleep(0.08)
            save_all()
            fetch_details_for_new()

    # Phase 5: Common food terms
    print(f"\n=== Phase 5: Food term queries ===")
    food_terms = [
        "oil", "flour", "sugar", "salt", "milk", "cream", "butter", "cheese",
        "chicken", "beef", "lamb", "fish", "shrimp", "salmon", "tuna",
        "rice", "wheat", "corn", "oat", "barley", "rye", "quinoa",
        "tomato", "onion", "garlic", "pepper", "potato", "carrot", "spinach",
        "apple", "banana", "orange", "lemon", "grape", "berry", "mango",
        "water", "juice", "vinegar", "sauce", "paste", "powder", "extract",
        "egg", "honey", "chocolate", "cocoa", "vanilla", "cinnamon", "ginger",
        "bean", "lentil", "pea", "nut", "almond", "walnut", "cashew",
        "yogurt", "whey", "protein", "fiber", "starch", "gelatin", "yeast",
        "soy", "tofu", "sesame", "coconut", "olive", "sunflower", "palm",
        "dried", "fresh", "frozen", "raw", "cooked", "roasted", "fried",
        "seed", "leaf", "root", "fruit", "vegetable", "meat", "poultry",
        "spice", "herb", "seasoning", "marinade", "dressing", "syrup",
        "bread", "noodle", "pasta", "cereal", "cracker", "biscuit",
        "date", "fig", "raisin", "prune", "apricot", "peach", "pear",
        "mushroom", "celery", "cucumber", "lettuce", "cabbage", "broccoli",
        "pumpkin", "squash", "eggplant", "zucchini", "avocado",
        "turkey", "duck", "goat", "veal", "liver", "kidney",
        "crab", "lobster", "oyster", "clam", "squid", "octopus",
        "margarine", "shortening", "lard", "ghee",
        "condensed", "evaporated", "skim", "whole",
        "baking", "cooking", "frying", "grilling",
        "supplement", "vitamin", "mineral", "calcium", "iron", "zinc",
        "preservative", "additive", "emulsifier", "stabilizer", "thickener",
        "caramel", "fondant", "marzipan", "meringue", "custard", "pudding",
    ]
    for term in food_terms:
        hits, nb = search(term)
        new = add_hits(hits)
        if new > 0:
            print(f"  '{term}': {len(hits)} hits, {new} new | Total: {len(all_ingredients)}")
        time.sleep(0.08)

    # Phase 6: Numeric
    print(f"\n=== Phase 6: Numeric queries ===")
    for n in range(10):
        hits, nb = search(str(n))
        new = add_hits(hits)
        if new > 0:
            print(f"  '{n}': {len(hits)} hits, {new} new | Total: {len(all_ingredients)}")
        time.sleep(0.1)

    save_all()
    save_search_progress(99, "done")
    fetch_details_for_new()
    print(f"\n=== ALL DONE: {len(all_ingredients)} ingredients, {sum(1 for v in details.values() if v)} with details ===")

# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def export_to_excel():
    # Load data
    if os.path.exists(DETAILS_FILE):
        with open(DETAILS_FILE) as f:
            data = json.load(f)
    else:
        print("No details file found. Run scraper first.")
        return

    valid = [d for d in data.values() if d is not None]
    if not valid:
        print("No ingredient data to export.")
        return

    print(f"\nExporting {len(valid)} ingredients to Excel...")

    # Collect nutrient names sorted by display order
    nutrient_order = {}
    for d in valid:
        for n in (d.get("nutrients") or []):
            name = n["name"]
            if name not in nutrient_order:
                nutrient_order[name] = n.get("order", 999)
    sorted_nutrients = sorted(nutrient_order.keys(), key=lambda x: nutrient_order[x])

    # Collect allergen names
    allergen_names = set()
    for d in valid:
        for a in (d.get("allergens") or []):
            allergen_names.add(a.get("name", ""))
    sorted_allergens = sorted(allergen_names)

    base_cols = ["id", "name", "name_ar", "friendly_name", "friendly_name_ar",
                 "data_type", "dataset", "category", "serving_size", "brand_name", "brand_owner"]
    all_headers = base_cols + sorted_nutrients + sorted_allergens

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ingredients"

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for ci, name in enumerate(all_headers, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = border

    valid.sort(key=lambda x: (x.get("name") or "").lower())
    for ri, ing in enumerate(valid, 2):
        cat = ing.get("category") or {}
        row = [
            ing.get("id"), ing.get("name"), ing.get("name_ar"),
            ing.get("friendly_name"), ing.get("friendly_name_ar"),
            ing.get("data_type"), ing.get("dataset"), cat.get("name", ""),
            ing.get("serving_size"), ing.get("brand_name"), ing.get("brand_owner"),
        ]
        nmap = {}
        for n in (ing.get("nutrients") or []):
            nmap[n["name"]] = (n.get("pivot") or {}).get("amount")
        row += [nmap.get(n) for n in sorted_nutrients]

        aset = {a.get("name", "") for a in (ing.get("allergens") or [])}
        row += ["Yes" if a in aset else "" for a in sorted_allergens]

        for ci, val in enumerate(row, 1):
            if isinstance(val, (dict, list)):
                val = json.dumps(val, ensure_ascii=False)
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = border

    for ci in range(1, len(all_headers) + 1):
        ml = len(str(ws.cell(row=1, column=ci).value or ""))
        for row in ws.iter_rows(min_row=2, max_row=min(50, ws.max_row), min_col=ci, max_col=ci):
            for cell in row:
                if cell.value is not None:
                    ml = max(ml, min(len(str(cell.value)), 40))
        ws.column_dimensions[get_column_letter(ci)].width = min(ml + 2, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_FILE)

    print(f"\nSaved: {EXCEL_FILE}")
    print(f"  {len(valid)} rows")
    print(f"  {len(base_cols)} base + {len(sorted_nutrients)} nutrient + {len(sorted_allergens)} allergen columns")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    if "--excel" in sys.argv:
        export_to_excel()
    else:
        scrape()
        export_to_excel()
